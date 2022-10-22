import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb1 = xl.load_workbook(filename)
    sheet1 = wb1['Sheet1']
    cell1 = sheet1['a1']# sheet1.cell(1,1)
    print(cell1)
    print(cell1.value)
    for row in range(1, sheet1.max_row+1):
        print(row)
    for row in range(2, sheet1.max_row+1):
        cell = sheet1.cell(row,3)
        print(cell.value)
        corrected_price = cell.value*0.9
        corrected_price_xl = sheet1.cell(row,4)
        corrected_price_xl.value = corrected_price

    values = Reference(sheet1,
              min_row=2,
              max_row=sheet1.max_row,
              min_col=4,
              max_col=4,)
    chart = BarChart()
    chart.add_data(values)
    sheet1.add_chart(chart,'e2')
    wb1.save(f'2{filename}')